"""구글시트를 저장소로 사용하는 모듈.

시트 구조:
  A: 이름  B: 주차  C~K: 10개 필드  L: 제출시간
"""
import io
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timezone, timedelta
import streamlit as st
from team_config import MEMBER_NAMES

KST = timezone(timedelta(hours=9))

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

FIELD_KEYS = [
    "acquired_data",
    "research_done",
    "research_plan",
    "task_done",
    "task_plan",
    "smart_care_space_done",
    "smart_care_space_plan",
    "project_confirmation_1",
    "project_confirmation_2_done",
    "project_confirmation_2_plan",
    "research_meeting",
    "director_meeting",
    "mohw_weekly",
]

FIELD_LABELS_KR = {
    "acquired_data": "획득데이터",
    "research_done": "연구실적",
    "research_plan": "연구계획",
    "task_done": "업무실적",
    "task_plan": "업무계획",
    "smart_care_space_done": "스마트돌봄스페이스_실적",
    "smart_care_space_plan": "스마트돌봄스페이스_계획",
    "project_confirmation_1": "사업단공통확인사항1",
    "project_confirmation_2_done": "사업단공통확인사항2_실적",
    "project_confirmation_2_plan": "사업단공통확인사항2_계획",
    "research_meeting": "연구소회의자료",
    "director_meeting": "주요간부회의자료",
    "mohw_weekly": "보산진주간일정",
}

HEADER = ["이름", "주차"] + [FIELD_LABELS_KR[k] for k in FIELD_KEYS] + ["제출시간"]
COL_COUNT = len(HEADER)


@st.cache_resource
def _get_client():
    info = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    return gspread.authorize(creds)


@st.cache_resource
def _get_sheet():
    """워크시트 핸들을 세션 내에서 재사용 (매번 API 호출 방지)."""
    client = _get_client()
    sheet_id = st.secrets["sheet"]["id"]
    ss = client.open_by_key(sheet_id)
    try:
        ws = ss.worksheet("submissions")
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title="submissions", rows=500, cols=COL_COUNT + 2)
        ws.append_row(HEADER)
    return ws


def _ensure_header(ws):
    values = ws.row_values(1)
    if values != HEADER:
        end_col = chr(ord('A') + COL_COUNT - 1)
        ws.update(f"A1:{end_col}1", [HEADER])


def _row_to_dict(row: list) -> dict:
    padded = list(row) + [""] * (COL_COUNT - len(row))
    return dict(zip(HEADER, padded))


@st.cache_data(ttl=30)
def _fetch_all_values() -> list:
    """전체 시트를 한 번만 읽어서 30초간 캐시 (API 쿼터 절감)."""
    ws = _get_sheet()
    _ensure_header(ws)
    return ws.get_all_values()


def load_week(week: str) -> dict:
    records = _fetch_all_values()
    out = {}
    for r in records[1:]:
        row = _row_to_dict(r)
        if row.get("주차", "") == week:
            data = {k: row.get(FIELD_LABELS_KR[k], "") for k in FIELD_KEYS}
            data["submitted_at"] = row.get("제출시간", "")
            out[row["이름"]] = data
    return out


def latest_submission(name: str):
    """이름의 가장 최근(주차 내림차순) 제출 1건 → (주차, {필드키: 텍스트}) 또는 None.
    주차는 'YYYY-MM-DD' 문자열이라 문자열 비교 = 날짜순."""
    records = _fetch_all_values()
    best_week, best = None, None
    for r in records[1:]:
        row = _row_to_dict(r)
        if row.get("이름") != name:
            continue
        wk = row.get("주차", "")
        if not wk:
            continue
        if best_week is None or wk > best_week:
            best_week = wk
            best = {k: row.get(FIELD_LABELS_KR[k], "") for k in FIELD_KEYS}
    if best is None:
        return None
    return best_week, best


def save_submission(name: str, week: str, values: dict) -> str:
    """values = {필드키: 텍스트} — FIELD_KEYS의 부분 집합. 누락은 빈 문자열로 저장."""
    ws = _get_sheet()
    _ensure_header(ws)
    now = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
    field_values = [values.get(k, "") for k in FIELD_KEYS]
    new_row = [name, week] + field_values + [now]

    all_rows = ws.get_all_values()
    for i, r in enumerate(all_rows[1:], start=2):
        row = _row_to_dict(r)
        if row.get("이름") == name and row.get("주차", "") == week:
            end_col = chr(ord('A') + COL_COUNT - 1)
            ws.update(values=[new_row], range_name=f"A{i}:{end_col}{i}")
            _fetch_all_values.clear()  # 캐시 무효화
            return "updated"
    ws.append_row(new_row)
    _fetch_all_values.clear()
    return "created"


def build_full_backup_xlsx() -> bytes:
    """제출함 스프레드시트의 모든 탭을 엑셀 1개로 덤프(오프라인 백업용).

    submissions·구매요청·문서협업·장비현황·방문일지 등 모든 워크시트를 그대로
    각 시트로 저장. 읽기 전용이라 데이터에 영향 없음.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    ss = _get_client().open_by_key(st.secrets["sheet"]["id"])
    wb = Workbook()
    wb.remove(wb.active)
    for sh in ss.worksheets():
        title = sh.title[:31]  # 엑셀 시트명 31자 제한
        for bad in r'\/?*[]:':
            title = title.replace(bad, "_")
        ws = wb.create_sheet(title=title or "sheet")
        for row in sh.get_all_values():
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def weeks_with_counts() -> list:
    """제출 데이터가 있는 주차 목록 — [(주차, 제출자수), ...] 최신순(내림차순).

    과거 회의록 드롭다운용. 실제 기록이 있는 주차만 반환하므로 빈 주차나
    오타 주차를 고를 일이 없다.
    """
    records = _fetch_all_values()
    counts = {}
    for r in records[1:]:
        row = _row_to_dict(r)
        wk = row.get("주차", "").strip()
        name = row.get("이름", "").strip()
        if wk and name:
            counts.setdefault(wk, set()).add(name)
    pairs = [(wk, len(names)) for wk, names in counts.items()]
    return sorted(pairs, key=lambda x: x[0], reverse=True)


def submission_status(week: str) -> list:
    data = load_week(week)
    out = []
    for name in MEMBER_NAMES:
        r = data.get(name)
        out.append({
            "name": name,
            "submitted": r is not None,
            "submitted_at": r["submitted_at"] if r else "",
        })
    return out
