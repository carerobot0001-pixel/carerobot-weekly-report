"""사업단 공통확인사항(최혜민) — 표 4개 저장 + 취합본 표 채워 한글/엑셀 생성.

표: 본부과제 용역(분야·발주금액·비고, 최대 6행) / 자산구매(품명·수량·구매금액·비고,
최대 12행), 각각 실적/계획. 취합본의 '빈 표 템플릿'(개인정보 제거본)의 리프 표에
셀을 채워 한글을 만든다(검증된 replace_cell 방식). 엑셀 출력도 제공.
"""
import io
import re
import zipfile
import html
from pathlib import Path

import gspread
import streamlit as st

from sheets_store import _get_client, KST
from hwpx_exporter import replace_cell, _patch_zip_flag_bits
from datetime import datetime

COMMON_WS_TITLE = "공통확인사항"
COMMON_HEADER = ["종류", "구분", "내용1", "내용2", "내용3", "내용4"]
KEYS = ["용역_실적", "용역_계획", "자산_실적", "자산_계획"]
EXTRA_KEY = "기타_내용"
YONG_MAX = 6
ASSET_MAX = 12
HWPX_YONG_MAX = 5
HWPX_ASSET_MAX = 10
TEMPLATE = Path(__file__).resolve().parent.parent / "사업단_공통확인사항_템플릿.hwpx"
COMMON_BODY_CHARPR = "34"


@st.cache_resource
def _ws():
    ss = _get_client().open_by_key(st.secrets["sheet"]["id"])
    try:
        return ss.worksheet(COMMON_WS_TITLE)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=COMMON_WS_TITLE, rows=200,
                              cols=len(COMMON_HEADER))
        ws.append_row(COMMON_HEADER)
        return ws


@st.cache_data(ttl=30)
def load_common() -> dict:
    """저장된 4개 표를 dict로. 각 값은 항목 리스트(리스트의 리스트)."""
    vals = _ws().get_all_values()
    out = {k: [] for k in KEYS}
    out[EXTRA_KEY] = ""
    for r in vals[1:]:
        if not any(c.strip() for c in r):
            continue
        r = (list(r) + [""] * 6)[:6]
        key = f"{r[0]}_{r[1]}"
        if key in KEYS:
            item = r[2:6]
            if not _is_total_row(item):
                out[key].append(item)
        elif key == EXTRA_KEY:
            out[EXTRA_KEY] = r[2]
    return out


def save_common(tables: dict) -> None:
    """4개 표를 시트에 통째로 저장(덮어쓰기)."""
    ws = _ws()
    rows = []
    for key in KEYS:
        종류, 구분 = key.split("_")
        for item in tables.get(key, []):
            item = (list(item) + [""] * 4)[:4]
            if any(str(x).strip() for x in item) and not _is_total_row(item):
                rows.append([종류, 구분] + [str(x).strip() for x in item])
    extra = str(tables.get(EXTRA_KEY, "")).strip()
    if extra:
        rows.append(["기타", "내용", extra, "", "", ""])
    n = len(ws.get_all_values())
    if n > 1:
        ws.delete_rows(2, n)
    if rows:
        ws.append_rows(rows, value_input_option="USER_ENTERED")
    load_common.clear()


def _num(s) -> int:
    try:
        return int(re.sub(r"[^\d]", "", str(s)) or 0)
    except Exception:
        return 0


def _fmt(n: int) -> str:
    return f"{n:,}" if n else ""


def _is_total_row(item) -> bool:
    vals = [str(x).strip().replace(" ", "") for x in (list(item) + [""])]
    return vals[0] in ("합계", "총계")


def _data_rows(items):
    return [list(item) for item in items if not _is_total_row(item)]


def _preview_text(tables: dict) -> str:
    lines = ["<사업단 공통확인사항>"]

    def block(title, header, items, max_rows):
        items = _data_rows(items)
        lines.append(f"<{title}>")
        lines.append("<" + "><".join(header) + ">")
        for i, item in enumerate(items[:max_rows], 1):
            vals = [str(x).strip() for x in item]
            if any(vals):
                lines.append(f"<{i}><" + "><".join(vals) + ">")

    block("본부과제 용역 - 실적", ["분야", "발주금액", "비고"],
          tables.get("용역_실적", []), HWPX_YONG_MAX)
    block("본부과제 용역 - 계획", ["분야", "발주금액", "비고"],
          tables.get("용역_계획", []), HWPX_YONG_MAX)
    block("본부과제 자산구매 - 실적", ["품명", "수량", "구매금액", "비고"],
          tables.get("자산_실적", []), HWPX_ASSET_MAX)
    block("본부과제 자산구매 - 계획", ["품명", "수량", "구매금액", "비고"],
          tables.get("자산_계획", []), HWPX_ASSET_MAX)
    extra = str(tables.get(EXTRA_KEY, "")).strip()
    if extra:
        lines.append("<기타내용>")
        lines.extend(extra.splitlines())
    return "\r\n".join(lines) + "\r\n"


def _extra_hwpx_block(text: str) -> str:
    text = str(text).strip()
    if not text:
        return ""
    lines = [x.strip() for x in text.splitlines() if x.strip()]
    body = " / ".join(lines)
    return (
        '<hp:p id="2147483648" paraPrIDRef="25" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">'
        '<hp:run charPrIDRef="11"><hp:t>기타내용</hp:t></hp:run>'
        '<hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="1100" textheight="1100" '
        'baseline="935" spacing="220" horzpos="0" horzsize="78516" flags="393216"/></hp:linesegarray></hp:p>'
        '<hp:p id="2147483648" paraPrIDRef="26" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">'
        f'<hp:run charPrIDRef="13"><hp:t>{html.escape(body)}</hp:t></hp:run>'
        '<hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="900" textheight="900" '
        'baseline="765" spacing="272" horzpos="0" horzsize="77132" flags="393216"/></hp:linesegarray></hp:p>'
    )


def _extra_cell_text(text: str) -> str:
    text = str(text).strip()
    if not text:
        return ""
    lines = [x.rstrip() for x in text.splitlines() if x.strip()]
    return "\n".join(["<湲고??댁슜>"] + lines)


def _leaf_tables(xml):
    out = []
    for m in re.finditer(r'<hp:tbl\b', xml):
        s = m.start()
        e = xml.find('</hp:tbl>', s) + len('</hp:tbl>')
        seg = xml[s:e]
        if seg.count('<hp:tbl') == 1:
            out.append((s, e, seg))
    return out


def _fill_yong(seg, items):
    items = _data_rows(items)
    total = 0
    for i in range(HWPX_YONG_MAX):
        r = i + 1
        분야, 발주, 비고 = ((list(items[i]) + ["", "", ""])[:3]
                          if i < len(items) else ("", "", ""))
        seg = replace_cell(seg, 0, r, str(r) if str(분야).strip() else "")
        seg = replace_cell(seg, 1, r, str(분야), override_color_id=COMMON_BODY_CHARPR)
        seg = replace_cell(seg, 2, r, _fmt(_num(발주)) if str(발주).strip() else "",
                           override_color_id=COMMON_BODY_CHARPR)
        seg = replace_cell(seg, 3, r, str(비고), override_color_id=COMMON_BODY_CHARPR)
        total += _num(발주)
    seg = replace_cell(seg, 2, HWPX_YONG_MAX + 1, _fmt(total),
                       override_color_id=COMMON_BODY_CHARPR)
    return seg


def _fill_asset(seg, items):
    items = _data_rows(items)
    total = 0
    for i in range(HWPX_ASSET_MAX):
        r = i + 1
        품명, 수량, 구매, 비고 = ((list(items[i]) + ["", "", "", ""])[:4]
                              if i < len(items) else ("", "", "", ""))
        seg = replace_cell(seg, 0, r, str(r) if str(품명).strip() else "")
        seg = replace_cell(seg, 1, r, str(품명), override_color_id=COMMON_BODY_CHARPR)
        seg = replace_cell(seg, 2, r, str(수량), override_color_id=COMMON_BODY_CHARPR)
        seg = replace_cell(seg, 3, r, _fmt(_num(구매)) if str(구매).strip() else "",
                           override_color_id=COMMON_BODY_CHARPR)
        seg = replace_cell(seg, 4, r, str(비고), override_color_id=COMMON_BODY_CHARPR)
        total += _num(구매)
    seg = replace_cell(seg, 3, HWPX_ASSET_MAX + 1, _fmt(total),
                       override_color_id=COMMON_BODY_CHARPR)
    return seg


def build_common_hwpx(tables: dict) -> bytes:
    """빈 표 템플릿에 4개 표를 채워 한글(HWPX) 생성."""
    data = TEMPLATE.read_bytes()
    with zipfile.ZipFile(io.BytesIO(data), 'r') as zin:
        xml = zin.read('Contents/section0.xml').decode('utf-8')
        order = [i.filename for i in zin.infolist()]
        infos = {i.filename: i for i in zin.infolist()}
        files = {n: zin.read(n) for n in zin.namelist()}

    leaves = _leaf_tables(xml)
    yong = [t for t in leaves if '발주금액' in t[2] and '구매금액' not in t[2]]
    asset = [t for t in leaves if '구매금액' in t[2]]
    if len(yong) < 2 or len(asset) < 2:
        raise RuntimeError("템플릿에서 용역/자산 표(각 2개)를 찾지 못했습니다.")

    edits = [
        (yong[0][0], yong[0][1], _fill_yong(yong[0][2], tables.get("용역_실적", []))),
        (yong[1][0], yong[1][1], _fill_yong(yong[1][2], tables.get("용역_계획", []))),
        (asset[0][0], asset[0][1], _fill_asset(asset[0][2], tables.get("자산_실적", []))),
        (asset[1][0], asset[1][1], _fill_asset(asset[1][2], tables.get("자산_계획", []))),
    ]
    for s, e, new_seg in sorted(edits, key=lambda x: x[0], reverse=True):
        xml = xml[:s] + new_seg + xml[e:]
    extra_text = _extra_cell_text(tables.get(EXTRA_KEY, ""))
    if extra_text:
        for col in (4, 5):
            xml = replace_cell(xml, col, 25, extra_text)

    files['Contents/section0.xml'] = xml.encode('utf-8')
    if 'Preview/PrvText.txt' in files:
        files['Preview/PrvText.txt'] = _preview_text(tables).encode('utf-8')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as zout:
        for n in order:
            o = infos[n]
            zi = zipfile.ZipInfo(n, date_time=o.date_time)
            zi.compress_type = o.compress_type
            zi.external_attr = o.external_attr
            zi.create_system = o.create_system
            zi.create_version = o.create_version
            zi.extract_version = o.extract_version
            zi.flag_bits = o.flag_bits
            zi.extra = o.extra
            zout.writestr(zi, files[n])
    return _patch_zip_flag_bits(buf.getvalue(), infos)


def build_common_xlsx(tables: dict) -> bytes:
    """4개 표를 엑셀로 (첨부 양식 유사). 한글 붙여넣기/대체용."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

    thin = Side(style="thin", color="999999")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill("solid", fgColor="D9E1F2")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "사업단 공통확인사항"

    def block(title, header, items, is_asset):
        items = _data_rows(items)
        ws.append([title])
        ws.append(["연번"] + header)
        hr = ws.max_row
        for c in ws[hr]:
            c.font = Font(bold=True); c.fill = hf; c.border = bd; c.alignment = ctr
        total = 0
        for i, it in enumerate(items, 1):
            it = list(it)
            if is_asset:
                품명, 수량, 구매, 비고 = (it + ["", "", "", ""])[:4]
                ws.append([i, 품명, 수량, _num(구매), 비고])
                total += _num(구매)
            else:
                분야, 발주, 비고 = (it + ["", "", ""])[:3]
                ws.append([i, 분야, _num(발주), 비고])
                total += _num(발주)
        tot_row = (["합계", "", "", total, ""] if is_asset else ["합계", "", total, ""])
        ws.append(tot_row)
        for row in ws.iter_rows(min_row=hr, max_row=ws.max_row):
            for c in row:
                c.border = bd
        ws.append([])

    block("<본부과제 용역> — 실적", ["분야", "발주금액", "비고"],
          tables.get("용역_실적", []), False)
    block("<본부과제 용역> — 계획", ["분야", "발주금액", "비고"],
          tables.get("용역_계획", []), False)
    block("<본부과제 자산구매> — 실적", ["품명", "수량", "구매금액", "비고"],
          tables.get("자산_실적", []), True)
    block("<본부과제 자산구매> — 계획", ["품명", "수량", "구매금액", "비고"],
          tables.get("자산_계획", []), True)
    extra = str(tables.get(EXTRA_KEY, "")).strip()
    if extra:
        ws.append(["기타내용"])
        ws.append([extra])
        ws[ws.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in {"A": 6, "B": 40, "C": 14, "D": 14, "E": 16}.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
