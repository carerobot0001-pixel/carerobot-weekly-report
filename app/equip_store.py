"""장비(기기) 사용현황 — 제출함 스프레드시트의 '장비현황' 탭에 마스터 1개로 관리.

엑셀의 '연구별 시트 자동분류'는 앱에서 '연구 필터'로 대체(수식 불필요).
전체 목록을 data_editor로 편집 → 통째로 시트 데이터영역에 덮어쓰기(save_all_equipment).
컬럼이 바뀌면 EQUIP_HEADER만 맞추면 됨. _ws()가 헤더 자동 보정.
"""
import io
import gspread
import streamlit as st

from sheets_store import _get_client

EQUIP_WS_TITLE = "장비현황"
EQUIP_HEADER = ["기기명", "S/N/Device ID", "자산번호", "연구", "플랫폼",
                "관련 앱 계정", "피험자명", "기간", "비고"]


@st.cache_resource
def _ws():
    ss = _get_client().open_by_key(st.secrets["sheet"]["id"])
    try:
        ws = ss.worksheet(EQUIP_WS_TITLE)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=EQUIP_WS_TITLE, rows=500,
                              cols=len(EQUIP_HEADER))
        ws.append_row(EQUIP_HEADER)
        return ws
    if ws.col_count < len(EQUIP_HEADER):
        ws.add_cols(len(EQUIP_HEADER) - ws.col_count)
    if ws.row_values(1) != EQUIP_HEADER:
        end = gspread.utils.rowcol_to_a1(1, len(EQUIP_HEADER))
        ws.update(values=[EQUIP_HEADER], range_name=f"A1:{end}")
    return ws


@st.cache_data(ttl=60)
def equip_rows() -> list:
    vals = _ws().get_all_values()
    out = []
    for r in vals[1:]:
        if not any(c.strip() for c in r):
            continue
        out.append((list(r) + [""] * len(EQUIP_HEADER))[:len(EQUIP_HEADER)])
    return out


def save_all_equipment(rows: list) -> int:
    """전체 장비 목록을 시트 데이터영역에 덮어쓰기 (행 추가/수정/삭제 일괄 반영).

    A2부터 새 데이터로 갱신하고, 기존이 더 길었으면 남는 행을 비운다.
    (전체 덮어쓰기라 동시 편집 시 마지막 저장이 우선 — 소규모 운영에 적합.)
    """
    ws = _ws()
    n = len(rows)
    if rows:
        end = gspread.utils.rowcol_to_a1(1 + n, len(EQUIP_HEADER))
        ws.update(values=rows, range_name=f"A2:{end}",
                  value_input_option="USER_ENTERED")
    existing = len(ws.get_all_values())
    if existing > 1 + n:
        c1 = gspread.utils.rowcol_to_a1(2 + n, 1)
        c2 = gspread.utils.rowcol_to_a1(existing, len(EQUIP_HEADER))
        ws.batch_clear([f"{c1}:{c2}"])
    equip_rows.clear()
    return n


def build_equip_xlsx(rows: list, title: str = "전체") -> bytes:
    """장비현황(현재 보기)을 엑셀로 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9E1F2")

    wb = Workbook()
    ws = wb.active
    ws.title = "장비현황"
    ws.append([f"장비 사용현황 — {title} ({len(rows)}개)"])
    ws.append([])
    ws.append(["번호"] + EQUIP_HEADER)
    hr = ws.max_row
    for c in ws[hr]:
        c.font = Font(bold=True)
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, r in enumerate(rows, 1):
        ws.append([i] + list(r))
    for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row,
                            min_col=1, max_col=len(EQUIP_HEADER) + 1):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {"A": 5, "B": 22, "C": 26, "D": 9, "E": 16, "F": 15, "G": 24,
              "H": 22, "I": 16, "J": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
