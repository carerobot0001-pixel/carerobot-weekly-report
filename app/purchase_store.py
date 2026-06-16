"""구매요청 — 기존 제출함 스프레드시트의 '구매요청' 탭에 누적 + 엑셀 양식 생성.

별도 시트를 새로 만들지 않고, 업무보고 제출함과 같은 스프레드시트
(secrets [sheet] id)에 '구매요청' 워크시트(탭)를 추가해 쓴다.
그 스프레드시트엔 서비스 계정이 이미 편집자라 추가 공유·secrets가 필요 없다.

한 번의 '구매요청'은 품목 여러 개로 구성되며, 시트에는 품목 1개당 1행으로
저장하고 같은 요청ID로 묶는다. 컬럼 구조가 바뀌면 PURCHASE_HEADER만 맞춘다.
"""
import io
import gspread
import streamlit as st

from sheets_store import _get_client, KST
from datetime import datetime

PURCHASE_WS_TITLE = "구매요청"
PURCHASE_HEADER = ["요청ID", "요청일시", "요청자", "품명", "품목(상세)", "단가",
                   "수량", "합계", "구매사유", "비고(구매처)",
                   "진행상황", "처리일자", "처리자"]

STATUS_DONE = "구매완료"
STATUS_PENDING = "요청"
# 진행상황/처리일자/처리자 컬럼 위치 (1-based) — 완료 처리 시 이 칸만 수정
STATUS_COL = PURCHASE_HEADER.index("진행상황") + 1   # K
DONE_DATE_COL = PURCHASE_HEADER.index("처리일자") + 1  # L
PROCESSOR_COL = PURCHASE_HEADER.index("처리자") + 1    # M


class RequestNotFound(Exception):
    """해당 요청ID의 행을 시트에서 찾지 못함 (이미 삭제/변경됨)."""

# 엑셀 양식(첨부 구매요청서)과 동일한 컬럼 순서
XLSX_HEADER = ["연번", "품명", "품목", "단가", "수량", "합계", "구매사유", "비고(구매처)"]


@st.cache_resource
def _ws():
    """제출함 스프레드시트의 '구매요청' 탭 핸들 (없으면 생성, 헤더/열수 보정)."""
    ss = _get_client().open_by_key(st.secrets["sheet"]["id"])
    try:
        ws = ss.worksheet(PURCHASE_WS_TITLE)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=PURCHASE_WS_TITLE, rows=1000,
                              cols=len(PURCHASE_HEADER))
        ws.append_row(PURCHASE_HEADER)
        return ws
    # 상태 컬럼 추가 등으로 헤더가 바뀌면 자동 보정 (기존 데이터 행은 빈 칸으로 패딩)
    if ws.col_count < len(PURCHASE_HEADER):
        ws.add_cols(len(PURCHASE_HEADER) - ws.col_count)
    if ws.row_values(1) != PURCHASE_HEADER:
        end = gspread.utils.rowcol_to_a1(1, len(PURCHASE_HEADER))
        ws.update(values=[PURCHASE_HEADER], range_name=f"A1:{end}")
    return ws


@st.cache_data(ttl=60)
def purchase_rows() -> list:
    """저장된 품목 행 목록(헤더 제외) — 헤더 길이에 맞춰 패딩."""
    vals = _ws().get_all_values()
    out = []
    for r in vals[1:]:
        if not any(c.strip() for c in r):
            continue
        out.append((list(r) + [""] * len(PURCHASE_HEADER))[:len(PURCHASE_HEADER)])
    return out


def add_purchase(requester: str, reason: str, items: list) -> tuple:
    """items = [{'품명','품목','단가','수량','비고'}...] 저장.

    반환: (요청ID, 저장된 품목 수, 총액)
    """
    ws = _ws()
    now = datetime.now(KST)
    req_id = now.strftime("%Y%m%d-%H%M%S-") + requester
    ts = now.strftime("%Y-%m-%d %H:%M")
    rows, total = [], 0
    for it in items:
        price = int(it.get("단가") or 0)
        qty = int(it.get("수량") or 0)
        subtotal = price * qty
        total += subtotal
        rows.append([req_id, ts, requester, it.get("품명", ""), it.get("품목", ""),
                     price, qty, subtotal, reason, it.get("비고", ""),
                     STATUS_PENDING, "", ""])
    if rows:
        ws.append_rows(rows, value_input_option="USER_ENTERED")
        purchase_rows.clear()
    return req_id, len(rows), total


def resolve_purchase(req_id: str, processor: str, done_date: str) -> int:
    """요청ID에 해당하는 모든 품목 행을 '구매완료'로 변경 (진행상황/처리일자/처리자).

    구매요청 탭은 앱만 append하므로 행 위치가 안정적이지만, 안전하게 매번
    요청ID로 행을 다시 찾아 그 행들의 K/L/M 칸만 수정한다. 반환: 처리된 품목 수.
    """
    ws = _ws()
    vals = ws.get_all_values()
    idxs = [i for i, r in enumerate(vals, start=1)
            if i > 1 and r and r[0] == req_id]
    if not idxs:
        raise RequestNotFound(req_id)
    data = [{"range": f"{gspread.utils.rowcol_to_a1(i, STATUS_COL)}:"
                      f"{gspread.utils.rowcol_to_a1(i, PROCESSOR_COL)}",
             "values": [[STATUS_DONE, done_date, processor]]} for i in idxs]
    ws.batch_update(data, value_input_option="USER_ENTERED")
    purchase_rows.clear()
    return len(idxs)


def delete_purchase_request(req_id: str) -> int:
    """요청ID에 해당하는 모든 품목 행 삭제(선택 삭제). 반환: 삭제된 행 수."""
    ws = _ws()
    vals = ws.get_all_values()
    idxs = [i for i, r in enumerate(vals, start=1)
            if i > 1 and r and r[0] == req_id]
    for i in reversed(idxs):
        ws.delete_rows(i)
    purchase_rows.clear()
    return len(idxs)


def clear_all_purchases() -> int:
    """누적 리스트 전체 삭제(헤더 1행은 보존). 반환: 삭제된 행 수."""
    ws = _ws()
    n = len(ws.get_all_values())
    if n > 1:
        ws.delete_rows(2, n)
    purchase_rows.clear()
    return max(0, n - 1)


# 누적 리스트 엑셀: 요청자 열을 '비고' 옆에 배치 (취합 담당 요청사항)
LIST_XLSX_HEADER = ["연번", "요청일시", "품명", "품목(상세)", "단가", "수량", "합계",
                    "구매사유", "비고(구매처)", "요청자", "진행상황", "처리일자", "처리자"]


def build_purchase_list_xlsx(rows: list) -> bytes:
    """누적 구매요청 전체를 엑셀로 생성 (요청자 열은 비고 옆)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "구매요청 누적"
    ws.append(["구매요청 누적 리스트"])
    ws.append([f"생성일 {datetime.now(KST).strftime('%Y-%m-%d %H:%M')}  ·  "
               f"총 {len(rows)}개 품목"])
    ws.append([])
    ws.append(LIST_XLSX_HEADER)
    head_row = ws.max_row
    for c in ws[head_row]:
        c.font = Font(bold=True)
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total = 0
    for i, r in enumerate(rows, 1):
        # r: 요청ID,요청일시,요청자,품명,품목,단가,수량,합계,구매사유,비고,진행,처리일,처리자
        price = int(r[5] or 0)
        qty = int(r[6] or 0)
        sub = int(r[7] or 0)
        total += sub
        ws.append([i, r[1], r[3], r[4], price, qty, sub, r[8], r[9], r[2],
                   r[10], r[11], r[12]])
    ws.append(["합계", "", "", "", "", "", total, "", "", "", "", "", ""])

    for row in ws.iter_rows(min_row=head_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=len(LIST_XLSX_HEADER)):
        for c in row:
            c.border = border
            if c.column_letter in ("E", "F", "G"):  # 단가/수량/합계
                c.alignment = right
                c.number_format = "#,##0"
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    widths = {"A": 5, "B": 15, "C": 18, "D": 40, "E": 11, "F": 6, "G": 13,
              "H": 16, "I": 22, "J": 9, "K": 9, "L": 12, "M": 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_purchase_xlsx(requester: str, reason: str, items: list) -> bytes:
    """첨부 구매요청서와 같은 양식의 엑셀 파일(bytes) 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "구매요청서"

    # 제목 + 요청자/일자
    ws.append([f"구매요청서 ({requester})"])
    ws.append([f"요청일자: {datetime.now(KST).strftime('%Y-%m-%d')}"])
    ws.append([])
    ws.append(XLSX_HEADER)
    head_row = ws.max_row
    for c in ws[head_row]:
        c.font = Font(bold=True)
        c.fill = head_fill
        c.alignment = center
        c.border = border

    total = 0
    for i, it in enumerate(items, 1):
        price = int(it.get("단가") or 0)
        qty = int(it.get("수량") or 0)
        subtotal = price * qty
        total += subtotal
        ws.append([i, it.get("품명", ""), it.get("품목", ""), price, qty,
                   subtotal, reason, it.get("비고", "")])
    ws.append(["합계", "", "", "", "", total, "", ""])

    # 본문 스타일 (헤더 다음 행부터 끝까지)
    for row in ws.iter_rows(min_row=head_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=len(XLSX_HEADER)):
        for c in row:
            c.border = border
            if c.column_letter in ("D", "E", "F"):  # 단가/수량/합계
                c.alignment = right
                c.number_format = "#,##0"
            elif c.column_letter in ("C", "G", "H"):
                c.alignment = left
            else:
                c.alignment = center
    # 합계 행 강조
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    ws.cell(row=ws.max_row, column=1).alignment = center

    widths = {"A": 6, "B": 18, "C": 42, "D": 12, "E": 6, "F": 14, "G": 16, "H": 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
